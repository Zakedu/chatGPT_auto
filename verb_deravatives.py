import openpyxl
import nltk
from nltk.corpus import wordnet

with open('Word_10.txt', 'r') as f:
    text = f.read()

words = nltk.word_tokenize(text)

# 각 단어에 대한 파생어 정보를 저장할 리스트 생성
derivatives = {}

# 단어 별 파생어를 반복문을 통해 탐색 
for word in words:
    # 단어의 다양한 변화형 중 가능한 것들을 모두 찾기
    forms = set()
    for synset in wordnet.synsets(word):  #유의어
        for lemma in synset.lemmas():
            for form in lemma.derivationally_related_forms():
                forms.add(form.name())
            for form in lemma.pertainyms():
                forms.add(form.name())
        for form in synset.lemma_names():
            if form != word:
                forms.add(form)
        morphy_forms = wordnet.morphy(word)  #음소 별
        if morphy_forms is not None:
            forms.update(morphy_forms)

    # 파생어 사전에 단어와 그 변형 더해넣기
    derivatives[word] = list(forms)

wb = openpyxl.Workbook()
ws = wb.active

# 열 값 지정
ws['A1'] = 'Word'
ws['B1'] = 'Derivatives'

row = 2
for word, forms in derivatives.items():
    ws.cell(row=row, column=1, value=word)
    ws.cell(row=row, column=2, value=', '.join(forms))
    row += 1

wb.save('derivatives.xlsx')
