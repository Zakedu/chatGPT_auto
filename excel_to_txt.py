##vice versa excel to txt.

import openpyxl

# Load the Excel file
wb = openpyxl.load_workbook('/Users/jake/PycharmProjects/pythonProject/GitHUB/chatGPT_auto/성윤result.xlsx')
ws = wb.active

# Open the output file for writing
with open('output_.txt', 'w') as f:
    # Write each row to the file
    for row in ws.iter_rows(values_only=True):
        # Skip the header row
        if row[0] == '메모':
            continue
        # Write the row to the file
        for i, value in enumerate(row):
            if i == 0:
                f.write(f"[메모]\n")
            elif value:
                f.write(f"[{ws.cell(row=1, column=i+1).value}]\n{value}\n")
        f.write('\n')
