import os
import nltk
import openpyxl

# Set the path to the folder containing the text files
folder_path = '/Users/jake/PycharmProjects/pythonProject/GitHUB/tokenizer'

# Initialize the workbook
wb = openpyxl.Workbook()

# Loop through each text file in the folder
for filename in os.listdir(folder_path):
    if filename.endswith('.txt'):
        # Load the text file
        with open(os.path.join(folder_path, filename), 'r') as f:
            text = f.read()

        # Tokenize the text
        sentences = nltk.sent_tokenize(text)
        tokens = [sentence.split() for sentence in sentences]

        # Create a new worksheet for the text file
        ws = wb.create_sheet(title=filename)

        # Add the tokens to the worksheet
        for i, sentence in enumerate(sentences):
            for j, token in enumerate(tokens[i]):
                ws.cell(row=i+1, column=j+1, value=token)

# Remove the default sheet created by openpyxl
wb.remove(wb['Sheet'])

# Save the workbook to an Excel file
wb.save('output.xlsx')
